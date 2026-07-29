#!/usr/bin/env python3
"""Auto-commit structured expense rows into Expense_Tracker.xlsx — safely.

This is the write-side counterpart to export_expense_csv.py (the read-side
mirror). It is the ONLY sanctioned path that mutates the canonical book, and it
crosses the "daemon never auto-writes xlsx" guardrail deliberately, with Steve's
explicit approval, gated on:

  1. an expense-reviewer SHIP verdict   (the agent review Steve required)
  2. a mandatory timestamped backup     (02_Finance/Backups/ before any write)
  3. dedup vs the CSV mirror            (vendor+date+amount already filed → skip)
  4. formula-safety                     (append data rows only, never row >200,
                                         load data_only=False so summary formulas
                                         are preserved, not frozen)

RUN UNDER /usr/bin/python3 — that interpreter has openpyxl; the iris_studio
.venv does NOT. The shebang points there on purpose.

Input: a JSON file (or stdin) — a list of row objects, keys matching the 9
'Expense Log' columns:
    date, category, vendor, description, amount, recurring, paid_by,
    receipt_link, notes
`amount` is required and numeric; date/vendor required. The rest default to "".

Usage:
    commit_expense.py --rows rows.json --review path/to/_Review.md
    commit_expense.py --rows rows.json --review ... --dry-run
    commit_expense.py --selftest          # offline, temp workbook, no live file
"""
import argparse
import csv
import fcntl
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

VAULT = Path("/Users/steve/Documents/3SK/outputs")
DEFAULT_XLSX = VAULT / "02_Finance/Expense_Tracker.xlsx"
NOTIFY = Path("/Volumes/AI_Workspace/iris_studio/scripts/notify.sh")
SHEET = "Expense Log"
HEADER_ROW = 4              # row 4 holds the column headers
FIRST_DATA_ROW = 5
MAX_DATA_ROW = 200         # summary sheets aggregate 'Expense Log'!A5:A200 — past this, formulas miss the row
COLUMNS = ["date", "category", "vendor", "description", "amount",
           "recurring", "paid_by", "receipt_link", "notes"]


# ----------------------------------------------------------------------------- helpers
def _notify(msg: str) -> None:
    if NOTIFY.exists():
        try:
            subprocess.run([str(NOTIFY), msg], timeout=15, check=False)
        except Exception:
            pass


def _amount_key(val) -> str:
    """Normalize an amount to a stable dedup key ('40' and '40.00' collide)."""
    try:
        return f"{float(str(val).replace('$', '').replace(',', '').strip()):.2f}"
    except (TypeError, ValueError):
        return str(val).strip()


def _norm_date(val) -> str:
    """Canonicalize a date to YYYY-MM-DD so '2026-6-25' and '2026-06-25' dedup.

    openpyxl may hand back a datetime (xlsx date cell) or a string; agent input is
    a string. Unparseable values are returned stripped, unchanged (never guess)."""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def validate_rows(rows):
    """Coerce + validate input rows. Raises ValueError on the first bad row."""
    clean = []
    for i, r in enumerate(rows):
        if not isinstance(r, dict):
            raise ValueError(f"row {i}: not an object")
        date = str(r.get("date", "")).strip()
        vendor = str(r.get("vendor", "")).strip()
        if not date:
            raise ValueError(f"row {i}: missing 'date'")
        if not vendor:
            raise ValueError(f"row {i}: missing 'vendor'")
        if "amount" not in r or str(r.get("amount")).strip() == "":
            raise ValueError(f"row {i}: missing 'amount'")
        try:
            amount = float(str(r["amount"]).replace("$", "").replace(",", "").strip())
        except (TypeError, ValueError):
            raise ValueError(f"row {i}: amount {r['amount']!r} is not numeric")
        clean.append({
            "date": _norm_date(date), "category": str(r.get("category", "")).strip(),
            "vendor": vendor, "description": str(r.get("description", "")).strip(),
            "amount": amount, "recurring": str(r.get("recurring", "")).strip(),
            "paid_by": str(r.get("paid_by", "")).strip(),
            "receipt_link": str(r.get("receipt_link", "")).strip(),
            "notes": str(r.get("notes", "")).strip(),
        })
    return clean


# Gmail message ids are server-assigned hex strings (16 chars in every corpus
# example seen so far); widened to 6-32 so a shorter/longer real id still matches
# rather than silently falling back to the weaker key.
_MSGID_RE = re.compile(r"Gmail msg-id[:\s]+([0-9a-fA-F]{6,32})")


def _msg_id(notes):
    """The Gmail msg-id embedded in a `notes` field, lowercased, or None."""
    m = _MSGID_RE.search(str(notes or ""))
    return m.group(1).lower() if m else None


def _dupe_report_line(row):
    """One human-readable line for a DROPPED row — names which row and via which
    key it matched, so a drop reads as evidence instead of a bare count."""
    mid = _msg_id(row.get("notes"))
    via = f"msg-id {mid}" if mid else "date+vendor+amount"
    return f"  {row['date']}  {row['vendor']}  ${row['amount']:.2f}  (dup via {via})"


def _dedup_key(date, vendor, amount, notes=None):
    """PRIMARY key: (Gmail msg-id, amount) when a msg-id is present in `notes`.
    Falls back to the (date, vendor, amount) triple only when no msg-id is found.

    CORRECTED 2026-07-29 (round-4 review) — the original justification here was
    self-contradicting and its evidence disproved it. It claimed date/vendor
    "DO drift" on a re-draft, then cited three sidecars in which date/vendor/
    amount are byte-identical. Re-pulled: only category/description/paid_by
    drift, and NONE of those is in the triple, so the plain triple already
    deduped the very case msg-id was introduced to fix. Corpus-wide the msg-id
    key changes nothing on its own (34 book rows + every draft sidecar: triple
    collisions = 0, msg-ids with >1 distinct amount = 0).

    It is kept because it closes a DIFFERENT hole the triple cannot: if vendor
    or date ever does drift between re-drafts of one receipt, the triple stops
    matching and files a duplicate. The server-assigned msg-id never drifts. See
    DedupIndex.has — the two keys are unioned, never ranked, because each has a
    duplicate-money path the other closes.

    Why msg-id ALONE is not enough (round-2 audit, 2026-07-28 REGRESSION): one
    Gmail message can carry TWO distinct line-item charges (a consolidated
    cloud bill, an invoice with two SKUs) — msg-id-only silently dropped the
    second amount as a "duplicate" of the first. Adding amount into the msgid
    key keeps both real cases: three re-drafts of the SAME charge (same
    msg-id + same amount, drifting category/description/paid_by) -> one row;
    two DIFFERENT charges in one message (same msg-id, different amount) ->
    two rows.

    ponytail: the residual ceiling is two GENUINELY IDENTICAL amounts inside
    one message (e.g. two separate $9.99 line items) -- those still collapse
    to one row. No such case exists in the live corpus (checked: every
    multi-charge message seen has distinct amounts). If one ever appears, the
    dupe report (_dupe_report_line) at least NAMES it as a msg-id+amount
    match so a human can catch it, rather than vanishing into a bare count."""
    mid = _msg_id(notes)
    if mid:
        return ("msgid", mid, _amount_key(amount))
    return ("triple", _norm_date(date), str(vendor).strip().lower(), _amount_key(amount))


class DedupIndex:
    """Dedup history that survives a ONE-SIDED msg-id.

    ROUND-3 REGRESSION (2026-07-28 review). `_dedup_key` returns EITHER a
    ("msgid", ...) key OR a ("triple", ...) key — never both — so the two sides of
    a comparison could key on different things and never match:

        history row WITHOUT a msg-id -> ('triple', '2026-05-18', 'anthropic', '100.00')
        incoming row WITH one        -> ('msgid', '19f2afe49b43b12b', '100.00')
        -> no match -> a DUPLICATE $100 row committed to the Schedule C book.

    That is not an edge case: **21 of 34 rows in the live Expense_Tracker.xlsx
    carry no msg-id** (62%), and `routines/expense-autocommit.prompt` runs this
    unattended every day asserting "idempotent (dedup), so a re-run never
    double-writes."

    Precedence, not a plain union — a union would re-break what the msg-id key was
    introduced to fix (two genuinely distinct same-day/same-vendor/same-amount
    charges collapsing into one):

      * incoming HAS a msg-id  -> match on msgid; fall back to the triple ONLY
        against history rows that have no msg-id (the one-sided case). Two
        mid-bearing rows with distinct mids therefore still stay two rows.
      * incoming has NO msg-id -> match on the triple against ALL history, since
        that is the only key both sides can share.
    """

    def __init__(self):
        # `mid_keys` may contain None (a row with no msg-id contributes one). That
        # is deliberate and harmless: has() only reaches the membership test when
        # its OWN mid_key is truthy, so a None entry can never be matched. Storing
        # it unconditionally removes the branch that used to sort rows into a
        # `midless_triples` set — that set was left WRITE-ONLY when the union
        # replaced the precedence rule, and its now-dead if/else survived mutation
        # precisely because nothing could observe which side ran. Deleting the dead
        # state removes the mutant rather than papering over it with a fixture that
        # could not kill it anyway. (Round-4 review follow-up, 2026-07-29.)
        self.mid_keys = set()     # msgid keys; may include None, never matched
        self.all_triples = set()  # every row's triple

    @staticmethod
    def _parts(date, vendor, amount, notes=None):
        mid = _msg_id(notes)
        triple = ("triple", _norm_date(date), str(vendor).strip().lower(),
                  _amount_key(amount))
        mid_key = ("msgid", mid, _amount_key(amount)) if mid else None
        return mid_key, triple

    def add(self, date, vendor, amount, notes=None):
        mid_key, triple = self._parts(date, vendor, amount, notes)
        self.all_triples.add(triple)
        self.mid_keys.add(mid_key)

    def has(self, date, vendor, amount, notes=None):
        """Already filed if EITHER key matches. Union, not precedence.

        ROUND-4 (2026-07-29 review). This used to let the msg-id key WIN when both
        sides carried one, falling back to the triple only for a one-sided msg-id.
        That was the wrong winner: BOTH keys have a duplicate-money path, and
        precedence closes one while opening the other.

          * vendor/date drift on a re-draft defeats the TRIPLE  -> msg-id catches it
          * one charge arriving as TWO Gmail messages (invoice + receipt, different
            server-assigned msg-ids, identical date/vendor/amount) defeats the
            MSG-ID -> the triple catches it

        Precedence took the second case from 1 row to 2 — reproduced end-to-end:
        HEAD filed $100.00 once, the precedence build filed it twice for $200.00,
        in a script `routines/expense-autocommit.prompt` runs unattended daily.
        Over-filing a Schedule C book is the worse direction, and invoice+receipt
        is an ordinary vendor behaviour, not an edge case.

        ponytail: the residual ceiling is the mirror case — two GENUINELY distinct
        charges sharing date+vendor+amount but carrying different msg-ids, which
        the triple collapses to one. There are ZERO such rows in the live corpus
        (checked: 34 book rows + every draft sidecar; same-triple-different-msgid
        = 0), and _dupe_report_line names which key matched, so a human sees the
        reason rather than a bare count. Revisit only if such a row appears."""
        mid_key, triple = self._parts(date, vendor, amount, notes)
        if mid_key and mid_key in self.mid_keys:
            return True
        return triple in self.all_triples

    def __len__(self):
        return len(self.all_triples)


def load_dedup_keys(xlsx_path: Path):
    """Set of (date, vendor_lower, amount_key) already filed.

    Reads the xlsx 'Expense Log' itself (authoritative) — NOT the CSV mirror,
    which can silently lag the book if a prior CSV regen failed. Unions in the
    CSV too as belt-and-suspenders, but the xlsx is the source of truth."""
    keys = DedupIndex()
    if xlsx_path.exists():
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        if SHEET in wb.sheetnames:
            ws = wb[SHEET]
            # max_col=9, not 5: column 9 (Notes) is where the msg-id primary key
            # lives — reading only through 'amount' silently dropped it and the
            # xlsx path fell back to the weaker triple for every historical row.
            # mutequiv: max_col=9->10 is unobservable — row[:5] and row[8] below
            # are fixed-index slices that never read position 9, so a wider fetch
            # changes nothing any caller can see. Proven by running the full
            # msg-id fixture suite (which DOES notice an under-read, e.g. max_col
            # dropped to 8 loses the Notes column and fails the round-trip
            # assertion below) against this mutant: it stays green. The value
            # must be >= 9 (a real constraint, tested); "== 9 exactly" is not.
            for row in ws.iter_rows(min_row=FIRST_DATA_ROW, max_col=9, values_only=True):
                date, _cat, vendor, _desc, amount = row[:5]
                notes = row[8] if len(row) > 8 else None
                if date is None and vendor is None and amount is None:
                    continue
                keys.add(date, vendor, amount, notes)
        wb.close()
    csv_path = xlsx_path.with_suffix(".csv")
    if csv_path.exists():
        with open(csv_path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                keys.add(row.get("Date", ""), row.get("Vendor", ""),
                         row.get("Amount ($)", ""), row.get("Notes", ""))
    return keys


def review_is_ship(review_path: Path) -> bool:
    """True only if the leading YAML frontmatter says exactly status: ship.

    Parses ONLY the first ---...--- frontmatter block (so a 'status: ship' in the
    review body prose can't flip the gate), and requires the value to equal 'ship'
    exactly — 'ship-blocked' / 'ship-with-fixes' are NOT ship (binary gate)."""
    if not review_path or not review_path.exists():
        return False
    text = review_path.read_text(encoding="utf-8", errors="replace")
    m = re.match(r"\s*---\s*\n(.*?)\n---\s*(?:\n|$)", text, re.DOTALL)
    if not m:
        return False
    fm = m.group(1)
    sm = re.search(r"^status:\s*([^\s#]+)\s*$", fm, re.MULTILINE)
    return bool(sm and sm.group(1).strip().lower() == "ship")


def append_rows(xlsx_path: Path, rows):
    """Append validated rows to the Expense Log, formula-safe. Returns row numbers.

    Loads data_only=False so summary-sheet formulas survive the save; finds the
    first fully-blank data row and writes there; aborts before MAX_DATA_ROW so a
    row never lands outside the summary aggregation range.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    if SHEET not in wb.sheetnames:
        raise KeyError(f"'{SHEET}' not found in {xlsx_path}")
    ws = wb[SHEET]
    ncols = len(COLUMNS)

    r = FIRST_DATA_ROW
    while any(ws.cell(row=r, column=c).value is not None for c in range(1, ncols + 1)):
        r += 1

    if r + len(rows) - 1 > MAX_DATA_ROW:
        raise RuntimeError(
            f"append would reach row {r + len(rows) - 1} > {MAX_DATA_ROW} "
            f"(outside summary-formula range A5:A200). Extend the formulas first."
        )

    written = []
    for row in rows:
        ws.cell(row=r, column=1, value=row["date"])
        ws.cell(row=r, column=2, value=row["category"])
        ws.cell(row=r, column=3, value=row["vendor"])
        ws.cell(row=r, column=4, value=row["description"])
        ws.cell(row=r, column=5, value=row["amount"])
        ws.cell(row=r, column=6, value=row["recurring"])
        ws.cell(row=r, column=7, value=row["paid_by"])
        ws.cell(row=r, column=8, value=row["receipt_link"])
        ws.cell(row=r, column=9, value=row["notes"])
        written.append(r)
        r += 1

    # Atomic save: temp in same dir, then os.replace.
    fd, tmp = tempfile.mkstemp(dir=str(xlsx_path.parent), suffix=".xlsx.tmp")
    os.close(fd)
    try:
        wb.save(tmp)
        os.replace(tmp, xlsx_path)
    except BaseException:
        if os.path.exists(tmp):
            os.unlink(tmp)
        raise
    return written


def commit(rows, xlsx_path=DEFAULT_XLSX, review_path=None, force=False, dry_run=False):
    """Full pipeline: validate → lock → dedup → review gate → backup → append → CSV → notify.

    The dedup→append→save critical section runs under an exclusive flock so two
    concurrent runs (e.g. a manual fire racing the hourly sweep) can't both read
    the same first-empty row and clobber each other's write."""
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"tracker not found at {xlsx_path}")
    is_live = xlsx_path.resolve() == DEFAULT_XLSX.resolve()

    def notify(msg):  # only ping the real Telegram channel for the live book
        if is_live:
            _notify(msg)

    clean = validate_rows(rows)

    lock_path = xlsx_path.parent / f".{xlsx_path.name}.lock"
    with open(lock_path, "w") as lock:
        fcntl.flock(lock, fcntl.LOCK_EX)

        # Dedup vs the authoritative xlsx (not the lagging mirror).
        seen = load_dedup_keys(xlsx_path)
        fresh, dupes = [], []
        for row in clean:
            is_dupe = seen.has(row["date"], row["vendor"], row["amount"], row["notes"])
            (dupes if is_dupe else fresh).append(row)
            seen.add(row["date"], row["vendor"], row["amount"], row["notes"])

        # WHICH row was dropped, not just a count — a silent count is exactly the
        # class of failure the msg-id fix above exists to make loud (2026-07-27/28
        # two-pass audit): a dupe drop that was actually a wrongly-keyed distinct
        # row would previously vanish into "N duplicate(s) skipped" with no way to
        # tell which N or why.
        dupe_lines = "\n".join(_dupe_report_line(r) for r in dupes)

        if not fresh:
            msg = (f"commit_expense: nothing to commit ({len(dupes)} duplicate(s) skipped)."
                   + (f"\n{dupe_lines}" if dupes else ""))
            print(msg)
            # NO Telegram ping here. Round-2 added one, reasoning that an all-dupe
            # batch is "a money-affecting drop". It is not: the money is already in
            # the book — that is what made the rows duplicates. An all-dupe batch is
            # the DOCUMENTED healthy outcome of a re-run
            # (routines/expense-autocommit.prompt PASS 3: "The committer is
            # idempotent (dedup), so a re-run never double-writes"), and
            # com.iris.claude-code-retry replays jobs every 30 min, so this would
            # ping on ordinary operation. That breaks the standing alert-channel
            # rule — FAILURES ONLY, no routine pings — and a channel that cries wolf
            # on healthy runs is how a real alert gets ignored. The dupe report is
            # printed above and lands in the job log, which is the right surface for
            # an expected no-op. Round-4 review, 2026-07-29.
            return {"committed": [], "dupes": dupes, "rows": []}

        # Review gate.
        if not force and not review_is_ship(Path(review_path) if review_path else None):
            raise PermissionError(
                "expense-reviewer SHIP verdict required "
                f"(--review pointing at frontmatter status: ship), or --force. Got: {review_path}"
            )
        if force:  # an unlogged bypass of the accounting-review control is the gap, so log it
            notify(f"⚠️ commit_expense: review gate FORCED (no SHIP verdict) on {len(fresh)} row(s).")

        if dry_run:
            print(f"[dry-run] would commit {len(fresh)} row(s), skip {len(dupes)} dupe(s):")
            for row in fresh:
                print(f"  {row['date']}  {row['vendor']}  ${row['amount']:.2f}  {row['category']}")
            if dupes:
                print(f"[dry-run] dupe(s) that would be skipped:\n{dupe_lines}")
            return {"committed": [], "dupes": dupes, "rows": fresh, "dry_run": True}

        # Mandatory backup BEFORE any write — into the book's own Backups dir.
        backup_dir = xlsx_path.parent / "Backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H%M%SZ")
        backup = backup_dir / f"Expense_Tracker_pre-autocommit_{stamp}.xlsx"
        shutil.copy2(xlsx_path, backup)

        try:
            written = append_rows(xlsx_path, fresh)
        except BaseException as e:
            notify(f"commit_expense FAILED on append: {e} (backup safe at {backup.name})")
            raise

        # Regenerate the CSV mirror so dedup stays current.
        try:
            sys.path.insert(0, str(Path(__file__).resolve().parent))
            import export_expense_csv
            export_expense_csv.export(xlsx_path)
        except Exception as e:
            notify(f"commit_expense: rows written (xlsx rows {written}) but CSV regen FAILED: {e}")

    total = sum(r["amount"] for r in fresh)
    lines = "\n".join(f"  {r['date']}  {r['vendor']}  ${r['amount']:.2f}" for r in fresh)
    dupe_note = f"\n{len(dupes)} dupe(s) skipped:\n{dupe_lines}" if dupes else ""
    notify(f"💸 Expense auto-commit: {len(fresh)} row(s), ${total:.2f} → rows {written}\n{lines}{dupe_note}\n(backup: {backup.name})")
    print(f"Committed {len(fresh)} row(s) to rows {written}; {len(dupes)} dupe(s) skipped."
          + (f"\n{dupe_lines}" if dupes else "") + f" Backup: {backup}")
    return {"committed": written, "dupes": dupes, "rows": fresh, "backup": str(backup)}


# ----------------------------------------------------------------------------- self-test
def _selftest():
    """Offline: build a synthetic workbook with the real structure, exercise append
    + dedup + the row-ceiling guard against a TEMP file. Never touches the live book."""
    import tempfile as _tf
    with _tf.TemporaryDirectory() as d:
        xlsx = Path(d) / "Expense_Tracker.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET
        for c, name in enumerate(["Date", "Category", "Vendor", "Description",
                                  "Amount ($)", "Recurring?", "Paid By",
                                  "Receipt Link", "Notes"], start=1):
            ws.cell(row=HEADER_ROW, column=c, value=name)
        wb.save(xlsx)

        # append two rows
        rows = validate_rows([
            {"date": "2026-06-25", "vendor": "OpenAI", "amount": "40", "category": "API"},
            {"date": "2026-06-25", "vendor": "ElevenLabs", "amount": 22.0, "category": "API"},
        ])
        written = append_rows(xlsx, rows)
        assert written == [5, 6], written

        # reload, confirm values landed and the amount is a real number
        wb2 = openpyxl.load_workbook(xlsx)[SHEET]
        assert wb2.cell(row=5, column=3).value == "OpenAI"
        assert abs(wb2.cell(row=6, column=5).value - 22.0) < 1e-9

        # next append starts at row 7 (first empty)
        written2 = append_rows(xlsx, validate_rows(
            [{"date": "2026-06-26", "vendor": "X", "amount": 1}]))
        assert written2 == [7], written2

        # amount-key dedup normalizes 40 == 40.00
        assert _amount_key("40") == _amount_key(40.00) == "40.00"

        # validation rejects a non-numeric amount and a missing vendor
        for bad in ([{"date": "d", "vendor": "v", "amount": "abc"}],
                    [{"date": "d", "amount": 1}]):
            try:
                validate_rows(bad)
                assert False, "should have raised"
            except ValueError:
                pass

        # date + amount keys normalize so 40==40.00 and 2026-6-25==2026-06-25
        assert _amount_key("$1,234.50") == "1234.50"
        assert _norm_date("2026-6-25") == _norm_date("2026-06-25") == "2026-06-25"
        assert _dedup_key("2026-6-25", "OpenAI", "40") == _dedup_key("2026-06-25", "openai", 40.0)

        # review gate (binary, frontmatter-only, exact 'ship'):
        rv = Path(d) / "r.md"
        rv.write_text("---\nstatus: revise\n---\n")
        assert review_is_ship(rv) is False
        rv.write_text("---\ntype: expense-review\nstatus: ship\n---\n")
        assert review_is_ship(rv) is True
        assert review_is_ship(Path(d) / "nope.md") is False
        # 'ship-blocked' / 'ship-with-fixes' must NOT pass (C1)
        rv.write_text("---\nstatus: ship-blocked\n---\n")
        assert review_is_ship(rv) is False
        rv.write_text("---\nstatus: ship with fixes\n---\n")
        assert review_is_ship(rv) is False
        # a 'status: SHIP' only in the BODY prose must NOT pass (C2)
        rv.write_text("---\nstatus: revise\n---\nbody text\nstatus: SHIP\n")
        assert review_is_ship(rv) is False

        # row-ceiling guard: filling to the cap then one more must raise
        big = Path(d) / "big.xlsx"
        wb3 = openpyxl.Workbook(); ws3 = wb3.active; ws3.title = SHEET
        for c, name in enumerate(["Date"] * 9, start=1):
            ws3.cell(row=HEADER_ROW, column=c, value=name)
        for rr in range(FIRST_DATA_ROW, MAX_DATA_ROW + 1):
            ws3.cell(row=rr, column=1, value="x")  # fill exactly to the cap
        wb3.save(big)
        try:
            append_rows(big, validate_rows([{"date": "d", "vendor": "v", "amount": 1}]))
            assert False, "ceiling guard should have raised"
        except RuntimeError:
            pass

        # C3: dedup keys come from the xlsx itself, even with NO csv mirror present.
        keys = load_dedup_keys(xlsx)  # xlsx has OpenAI/40, ElevenLabs/22, X/1
        assert not xlsx.with_suffix(".csv").exists()  # no mirror at all
        assert keys.has("2026-06-25", "OpenAI", "40.00")
        assert keys.has("2026-06-25", "ElevenLabs", 22)

        # D — msg-id primary-key dedup (2026-07-28 two-pass audit fix). Verified
        # against the live corpus: msg-id 19f2afe49b43b12b's $100 Anthropic row
        # appears in THREE Expense_Tracker_Drafts sidecars with category/
        # description/paid_by all drifting between drafts. A msg-id in `notes`
        # must dedup even when every OTHER triple field drifts; and — the
        # opposite direction, which the audit's rejected "widen the key" fix
        # would have gotten backwards — a genuinely DIFFERENT msg-id must NOT be
        # swallowed just because date+vendor+amount happen to collide.
        notes_a = "Filed via run AM; Gmail msg-id 19f2afe49b43b12b. gap month."
        notes_b = ("Filed via run AM (2026-07-06); Gmail msg-id 19f2afe49b43b12b. "
                   "resend, different wording.")
        notes_c = "Gmail msg-id aaaaaaaaaaaaaaaa. unrelated separate charge."
        assert _msg_id(notes_a) == "19f2afe49b43b12b"
        assert _msg_id("no id in this note at all") is None
        # SAME msg-id, drifted date/vendor/amount-string -> SAME key.
        assert (_dedup_key("2026-05-18", "Anthropic", "100", notes_a)
                == _dedup_key("2026-06-01", "Anthropic Inc.", "100.00", notes_b))
        # DIFFERENT msg-id, IDENTICAL triple -> DIFFERENT key (msg-id overrides
        # the triple rather than being ignored on a collision).
        assert (_dedup_key("2026-05-18", "Anthropic", "100", notes_a)
                != _dedup_key("2026-05-18", "Anthropic", "100", notes_c))
        # No msg-id anywhere -> unchanged triple-fallback behavior.
        assert (_dedup_key("2026-05-18", "Anthropic", "100")
                == _dedup_key("2026-05-18", "anthropic", "100.00"))
        # Pin the SHAPE too, not just equality between two no-msg-id calls (which
        # a forced `if True:` on the mid-check would satisfy identically, since
        # both sides would collapse to the same ("msgid", None) tuple).
        assert _dedup_key("2026-05-18", "Anthropic", "100") == \
            ("triple", "2026-05-18", "anthropic", "100.00")
        # Pin the msgid tuple's own shape too (3-tuple: tag, mid, amount_key).
        assert _dedup_key("2026-05-18", "Anthropic", "100", notes_a) == \
            ("msgid", "19f2afe49b43b12b", "100.00")

        # D2 — round-2 audit REGRESSION (2026-07-28): msg-id ALONE as the key
        # collapsed two DISTINCT charges sharing one Gmail message (a
        # consolidated cloud bill with two line items). Amount must be part of
        # the msgid key: same msg-id + DIFFERENT amount -> DIFFERENT key (must
        # NOT dedup); same msg-id + SAME amount (a drifted re-draft) -> SAME
        # key (must still dedup, D above).
        notes_two_charges = "Filed via run AM; Gmail msg-id cafe1234cafe5678. cloud bill."
        assert (_dedup_key("2026-07-15", "Google Cloud", "31.20", notes_two_charges)
                != _dedup_key("2026-07-15", "Google Cloud", "8.40", notes_two_charges)), \
            "two distinct amounts sharing one msg-id must NOT collapse to one key"
        # End-to-end reproduction of the exact reported repro: one batch, one
        # msg-id, two distinct amounts -> commit() must keep BOTH, not drop one
        # as a false "duplicate" of the other.
        two_xlsx = Path(d) / "TwoCharge_Tracker.xlsx"
        wb5 = openpyxl.Workbook(); ws5 = wb5.active; ws5.title = SHEET
        for c, name in enumerate(["Date", "Category", "Vendor", "Description",
                                  "Amount ($)", "Recurring?", "Paid By",
                                  "Receipt Link", "Notes"], start=1):
            ws5.cell(row=HEADER_ROW, column=c, value=name)
        # Pin the fixture's own column placement (same pattern as ws4/wse above) --
        # a shifted start= would otherwise be invisible here.
        assert ws5.cell(row=HEADER_ROW, column=1).value == "Date"
        assert ws5.cell(row=HEADER_ROW, column=9).value == "Notes"
        wb5.save(two_xlsx)
        result = commit(
            [{"date": "2026-07-15", "vendor": "Google Cloud", "amount": "31.20",
              "category": "Cloud", "notes": notes_two_charges},
             {"date": "2026-07-15", "vendor": "Google Cloud", "amount": "8.40",
              "category": "Cloud", "notes": notes_two_charges}],
            xlsx_path=two_xlsx, force=True, dry_run=True,
        )
        committed_amounts = sorted(r["amount"] for r in result["rows"])
        assert committed_amounts == [8.40, 31.20], (
            f"both distinct charges must survive batch-internal dedup, got {committed_amounts}")
        assert result["dupes"] == [], result["dupes"]

        # load_dedup_keys must actually READ column 9 (Notes) — it used to stop
        # at max_col=5 and silently drop every msg-id from history, so a drifted
        # re-draft of an already-filed row would NOT have deduped.
        mid_xlsx = Path(d) / "MsgId_Tracker.xlsx"
        wb4 = openpyxl.Workbook(); ws4 = wb4.active; ws4.title = SHEET
        for c, name in enumerate(["Date", "Category", "Vendor", "Description",
                                  "Amount ($)", "Recurring?", "Paid By",
                                  "Receipt Link", "Notes"], start=1):
            ws4.cell(row=HEADER_ROW, column=c, value=name)
        ws4.cell(row=FIRST_DATA_ROW, column=1, value="2026-05-18")
        ws4.cell(row=FIRST_DATA_ROW, column=3, value="Anthropic")
        ws4.cell(row=FIRST_DATA_ROW, column=5, value=100.0)
        ws4.cell(row=FIRST_DATA_ROW, column=9, value=notes_a)
        # Pin the fixture's own column placement directly (not just its downstream
        # effect) -- a shifted column write is otherwise invisible here because the
        # msg-id primary key makes date/vendor "don't care" once mid is present,
        # so a column-offset bug in the fixture writer would silently pass.
        assert ws4.cell(row=HEADER_ROW, column=1).value == "Date"
        assert ws4.cell(row=HEADER_ROW, column=9).value == "Notes"
        assert ws4.cell(row=FIRST_DATA_ROW, column=1).value == "2026-05-18"
        assert ws4.cell(row=FIRST_DATA_ROW, column=3).value == "Anthropic"
        assert ws4.cell(row=FIRST_DATA_ROW, column=5).value == 100.0
        wb4.save(mid_xlsx)
        mid_keys = load_dedup_keys(mid_xlsx)
        assert mid_keys.has("2026-06-01", "Anthropic Inc.", "100.00", notes_b), \
            "a drifted re-draft sharing the same msg-id must read as already-filed"

        # D2 — ONE-SIDED msg-id (round-3 review, 2026-07-28). 21 of 34 rows in the
        # live book carry NO msg-id, and expense-autocommit runs unattended daily
        # claiming idempotence. When only ONE side has a msg-id the two keys were
        # different SHAPES ('triple' vs 'msgid') and could never match, so the row
        # was committed AGAIN — a duplicate money row, silently.
        assert mid_keys.has("2026-05-18", "Anthropic", "100.00", ""), \
            "history HAS a msg-id, incoming has none -> must still dedup (one-sided)"
        _one_sided = DedupIndex()
        _one_sided.add("2026-05-18", "Anthropic", 100.0, "")          # history: no mid
        assert _one_sided.has("2026-05-18", "Anthropic", "100.00", notes_a), \
            "history has NO msg-id, incoming has one -> must still dedup (one-sided)"
        # D3 — the ACCEPTED CEILING of the union (assertion inverted 2026-07-29,
        # round-4 review). Under the earlier PRECEDENCE rule two mid-bearing rows
        # sharing a triple stayed separate; under the union they collapse. That is
        # a deliberate trade, not an oversight:
        #   * precedence filed ONE charge arriving as TWO Gmail messages (invoice
        #     + receipt, different server msg-ids, identical date/vendor/amount)
        #     as TWO rows — reproduced end-to-end at $200.00 vs HEAD's $100.00, in
        #     a script that runs unattended daily. Ordinary vendor behaviour.
        #   * the case it protected — two GENUINELY distinct charges sharing
        #     date+vendor+amount with different msg-ids — has ZERO instances in
        #     the live corpus (34 book rows + every draft sidecar).
        # Over-filing a Schedule C book is the worse direction, so the union wins
        # and this collapse is the documented ceiling. _dupe_report_line names
        # which key matched, so a human sees the reason, not a bare count.
        _both = DedupIndex()
        _both.add("2026-05-18", "Anthropic", 100.0, notes_a)
        assert _both.has("2026-05-18", "Anthropic", 100.0,
                         "Gmail msg-id ffffffffffffffff"), \
            ("same date+vendor+amount must dedup even with different msg-ids -- "
             "this is what stops one charge arriving as invoice+receipt being "
             "filed twice")

        # the drop report names WHICH row and via which key, not just a count.
        assert "msg-id 19f2afe49b43b12b" in _dupe_report_line(
            {"date": "2026-05-18", "vendor": "Anthropic", "amount": 100.0, "notes": notes_a})
        assert "date+vendor+amount" in _dupe_report_line(
            {"date": "2026-05-18", "vendor": "X", "amount": 1.0, "notes": ""})

        # E — commit() end-to-end, dry-run with a real duplicate present: the
        # dupe report must actually print (not just be computed and discarded).
        import io as _io
        import contextlib as _cl
        e_xlsx = Path(d) / "DryRun_Tracker.xlsx"
        wbe = openpyxl.Workbook(); wse = wbe.active; wse.title = SHEET
        for c, name in enumerate(["Date", "Category", "Vendor", "Description",
                                  "Amount ($)", "Recurring?", "Paid By",
                                  "Receipt Link", "Notes"], start=1):
            wse.cell(row=HEADER_ROW, column=c, value=name)
        assert wse.cell(row=HEADER_ROW, column=1).value == "Date"
        assert wse.cell(row=HEADER_ROW, column=9).value == "Notes"
        wbe.save(e_xlsx)
        append_rows(e_xlsx, validate_rows(
            [{"date": "2026-07-01", "vendor": "Adobe", "amount": "50", "category": "SaaS"}]))
        buf = _io.StringIO()
        with _cl.redirect_stdout(buf):
            commit(
                [{"date": "2026-07-02", "vendor": "Notion", "amount": "9", "category": "SaaS"},
                 {"date": "2026-07-01", "vendor": "Adobe", "amount": "50", "category": "SaaS"}],
                xlsx_path=e_xlsx, force=True, dry_run=True,
            )
        out = buf.getvalue()
        assert "dupe(s) that would be skipped" in out, out
        assert "Adobe" in out, out

        # The inverse: a dry-run with NO duplicates must NOT print the dupe-report
        # header at all -- kills the :317 if-test->True mutant, which would print
        # it (with an empty body) even when dupes is empty.
        buf2 = _io.StringIO()
        with _cl.redirect_stdout(buf2):
            commit(
                [{"date": "2026-07-03", "vendor": "Figma", "amount": "12", "category": "SaaS"}],
                xlsx_path=e_xlsx, force=True, dry_run=True,
            )
        out2 = buf2.getvalue()
        assert "dupe(s) that would be skipped" not in out2, out2

        # E2 — WITHIN-BATCH dedup (round-5 review F1). `seen.add(...)` in commit()'s
        # loop is a bare call statement, which mutation CANNOT see (the gate says so
        # in its own footer), and nothing covered it: deleting that line left all 43
        # suites green while one JSON batch containing the same charge twice filed
        # TWO rows — the exact duplicate-money outcome these rounds exist to stop,
        # in a script routines/expense-autocommit.prompt runs unattended daily.
        e2_xlsx = Path(d) / "WithinBatch_Tracker.xlsx"
        wb_e2 = openpyxl.Workbook(); ws_e2 = wb_e2.active; ws_e2.title = SHEET
        wb_e2.save(e2_xlsx)
        _dupe_row = {"date": "2026-08-01", "vendor": "Framer", "amount": "25",
                     "category": "SaaS"}
        e2 = commit([dict(_dupe_row), dict(_dupe_row)], xlsx_path=e2_xlsx,
                    force=True, dry_run=True)
        assert len(e2["rows"]) == 1, \
            f"the same charge twice in ONE batch must file once, got {len(e2['rows'])}"
        assert len(e2["dupes"]) == 1, \
            f"the second copy must be reported as a dupe, got {len(e2['dupes'])}"

        # E3 — the CSV-mirror msg-id path (round-5 review F2). The xlsx branch has a
        # dedicated max_col=9 fixture; its CSV twin had none, so replacing
        # row.get("Notes") with "" there left every suite green — silently killing
        # msg-id dedup for any row that reaches the index via the mirror.
        e3_xlsx = Path(d) / "CsvMirror_Tracker.xlsx"
        wb_e3 = openpyxl.Workbook(); ws_e3 = wb_e3.active; ws_e3.title = SHEET
        wb_e3.save(e3_xlsx)
        e3_csv = e3_xlsx.with_suffix(".csv")
        with open(e3_csv, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["Date", "Category", "Vendor", "Description", "Amount ($)",
                        "Recurring", "Paid By", "Receipt Link", "Notes"])
            w.writerow(["2026-05-18", "AI", "Anthropic", "credits", "100",
                        "", "", "", notes_a])
        e3_keys = load_dedup_keys(e3_xlsx)
        assert e3_keys.has("2026-06-01", "Anthropic Inc.", "100.00", notes_b), \
            ("a drifted re-draft must dedup against a msg-id that reached the index "
             "via the CSV MIRROR, not just the xlsx")

        # F — an all-dupes batch must NOT ping Telegram, on a real commit or a
        # dry-run. ASSERTION INVERTED 2026-07-29 (round-4 review): round-2 added
        # the ping calling an all-dupe batch "a money-affecting drop", and this
        # fixture pinned it. It is the opposite — the money is already in the book,
        # which is WHY the rows are duplicates, and an all-dupe batch is the
        # documented healthy outcome of the daily unattended re-run. Pinning a
        # routine ping violated the standing FAILURES-ONLY alert rule, so the
        # fixture was pinning the defect. DEFAULT_XLSX is still monkeypatched so a
        # regression that reintroduces the ping is actually caught here.
        global _notify, DEFAULT_XLSX
        _orig_notify, _orig_default_xlsx = _notify, DEFAULT_XLSX
        _pinged = []
        _notify = lambda msg: _pinged.append(msg)
        DEFAULT_XLSX = e_xlsx
        try:
            commit([{"date": "2026-07-01", "vendor": "Adobe", "amount": "50",
                    "category": "SaaS"}], xlsx_path=e_xlsx, force=True, dry_run=False)
            # Split, not `and`-combined: an `and`->`or` flip is unobservable when
            # both operands happen to already be True in this fixture.
            assert _pinged == [], (
                "an all-dupe batch is the healthy idempotent outcome of a re-run "
                f"and must not ping Telegram (FAILURES ONLY): {_pinged}")
            commit([{"date": "2026-07-01", "vendor": "Adobe", "amount": "50",
                    "category": "SaaS"}], xlsx_path=e_xlsx, force=True, dry_run=True)
            assert _pinged == [], f"a dry-run all-dupe batch must NOT notify: {_pinged}"
        finally:
            _notify, DEFAULT_XLSX = _orig_notify, _orig_default_xlsx

    print("commit_expense self-check: PASS")


# ----------------------------------------------------------------------------- cli
def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--rows", help="JSON file of row objects (or '-' for stdin)")
    ap.add_argument("--review", help="path to the expense-reviewer SHIP verdict")
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    ap.add_argument("--force", action="store_true", help="bypass the review gate (logged)")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--selftest", action="store_true")
    args = ap.parse_args()

    if args.selftest:
        _selftest()
        return

    if not args.rows:
        ap.error("--rows is required (JSON file or '-' for stdin)")
    try:
        raw = sys.stdin.read() if args.rows == "-" else Path(args.rows).read_text(encoding="utf-8")
        rows = json.loads(raw)
        if isinstance(rows, dict):
            rows = [rows]
        commit(rows, xlsx_path=args.xlsx, review_path=args.review,
               force=args.force, dry_run=args.dry_run)
    except Exception as e:
        _notify(f"commit_expense FAILED: {e}")
        sys.exit(f"ERROR: {e}")


if __name__ == "__main__":
    main()
