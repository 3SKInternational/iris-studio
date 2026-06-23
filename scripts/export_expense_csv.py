#!/usr/bin/env python3
"""Export the Expense_Tracker.xlsx 'Expense Log' sheet to a flat CSV ledger.

The .xlsx is binary and unreadable by the expense-categorizer subagent, which
broke receipt dedup (it could not confirm whether a charge was already filed).
This emits a machine-readable companion ledger the agent can grep before drafting
new rows. Run it after ANY edit to Expense_Tracker.xlsx to keep the CSV current.

Read-only on the xlsx; only ever writes the separate .csv (atomic swap).

Usage: python3 export_expense_csv.py [path/to/Expense_Tracker.xlsx]
Default path: 02_Finance/Expense_Tracker.xlsx in the 3SK vault.
"""
import csv
import os
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl

DEFAULT_XLSX = Path(
    "/Users/steve/Documents/3SK/outputs/02_Finance/Expense_Tracker.xlsx"
)
NOTIFY = Path("/Volumes/AI_Workspace/iris_studio/scripts/notify.sh")
SHEET = "Expense Log"


def _notify(msg: str) -> None:
    if NOTIFY.exists():
        try:
            subprocess.run([str(NOTIFY), msg], timeout=15, check=False)
        except Exception:
            pass


def export(xlsx_path: Path) -> Path:
    csv_path = xlsx_path.with_suffix(".csv")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET not in wb.sheetnames:
        raise KeyError(f"'{SHEET}' sheet not found in {xlsx_path}")
    ws = wb[SHEET]
    header = [c.value for c in ws[4]]  # row 4 is the column-header row
    ncols = len(header)
    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        row = row[:ncols]
        if all(c is None for c in row):  # fully blank row = end of data
            break
        rows.append(row)

    # Atomic write: temp file in same dir, then os.replace so a grep consumer
    # never sees a half-written ledger.
    fd, tmp = tempfile.mkstemp(dir=csv_path.parent, suffix=".csv.tmp")
    try:
        with os.fdopen(fd, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(header)
            w.writerows(rows)
        os.replace(tmp, csv_path)
    except BaseException:
        if os.path.exists(tmp):
            os.unlink(tmp)
        raise
    print(f"Wrote {csv_path} ({len(rows)} data rows).")
    return csv_path


if __name__ == "__main__":
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not path.exists():
        _notify(f"export_expense_csv: tracker not found at {path}")
        sys.exit(f"ERROR: tracker not found at {path}")
    try:
        export(path)
    except Exception as e:
        _notify(f"export_expense_csv FAILED: {e}")
        sys.exit(f"ERROR: {e}")
